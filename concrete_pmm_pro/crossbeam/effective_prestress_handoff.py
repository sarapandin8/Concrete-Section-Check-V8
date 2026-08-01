"""Source-gated Effective Prestress handoff for external FEA workflows.

The handoff exports the currently reviewed tendon stress/force chain without
pretending to calculate portal-frame secondary prestress inside Concrete
Section Pro.  External FEA remains responsible for primary/secondary response,
and verified SLS resultants return through the main Loads workspace.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import datetime, timezone
from io import BytesIO
import hashlib
import json
from math import isfinite
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EFFECTIVE_PRESTRESS_FEA_HANDOFF_SCHEMA = (
    "crossbeam-effective-prestress-fea-handoff-v2"
)
EFFECTIVE_PRESTRESS_FEA_HANDOFF_BASIS = (
    "Aps-weighted piecewise-trapezoidal integration over projected member station s"
)
FEA_ROUTE_DIRECT_EFFECTIVE_FORCE = "DIRECT_EFFECTIVE_FORCE"
FEA_ROUTE_JACKING_WITH_LOSSES = "JACKING_FORCE_WITH_FEA_LOSSES"
FEA_APPLICATION_ROUTES = (
    FEA_ROUTE_DIRECT_EFFECTIVE_FORCE,
    FEA_ROUTE_JACKING_WITH_LOSSES,
)
AUTOMATIC_HANDOFF_READY_STATUS = "SOURCE READY — FINAL-STAGE FEA HANDOFF"
TENDON_HANDOFF_EXPORT_COLUMNS = (
    "Tendon",
    "Aps (mm²)",
    "fpj (MPa)",
    "Pj (kN)",
    "Left Pe (kN)",
    "Mid Pe (kN)",
    "Right Pe (kN)",
    "Average fpe (MPa)",
    "Average Pe (kN)",
    "Average total loss (MPa)",
    "Average loss (% fpj)",
    "Remaining prestress (%)",
    "Source ID",
)


def _finite_float(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return float(default)
    return number if isfinite(number) else float(default)


def _records(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, pd.DataFrame):
        return [dict(row) for row in value.to_dict(orient="records")]
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return [dict(row) for row in value if isinstance(row, Mapping)]
    return []


def _collapse_tendon_station_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Collapse duplicate station faces once for one Tendon.

    Numeric values are averaged at a shared station.  Text labels are joined so
    the audit still records which point/face rows were represented.
    """

    grouped: dict[float, list[dict[str, Any]]] = {}
    for row in rows:
        station = round(_finite_float(row.get("Station s (m)")), 9)
        grouped.setdefault(station, []).append(row)

    numeric_fields = (
        "Aps (mm²)",
        "fpj (MPa)",
        "Friction (MPa)",
        "Anchorage (MPa)",
        "ES (MPa)",
        "Creep (MPa)",
        "Shrinkage (MPa)",
        "Relaxation (MPa)",
        "TD total (MPa)",
        "Total loss (MPa)",
        "Loss (% fpj)",
        "fpe preview (MPa)",
        "Pj (kN)",
        "Pe preview (kN)",
    )
    collapsed: list[dict[str, Any]] = []
    for station in sorted(grouped):
        station_rows = grouped[station]
        result: dict[str, Any] = {
            "Tendon": str(station_rows[0].get("Tendon") or ""),
            "Station s (m)": station,
            "Point": " / ".join(
                sorted(
                    {
                        str(row.get("Point") or "").strip()
                        for row in station_rows
                        if str(row.get("Point") or "").strip()
                    }
                )
            ),
            "Source rows collapsed": len(station_rows),
        }
        for field in numeric_fields:
            values = [_finite_float(row.get(field)) for row in station_rows]
            result[field] = sum(values) / len(values) if values else 0.0
        collapsed.append(result)
    return collapsed


def _trapezoidal_average(rows: list[dict[str, Any]], value_key: str) -> float | None:
    if not rows:
        return None
    points = sorted(
        (float(row["Station s (m)"]), _finite_float(row.get(value_key))) for row in rows
    )
    if len(points) == 1:
        return float(points[0][1])
    covered = float(points[-1][0]) - float(points[0][0])
    if covered <= 0.0:
        return float(points[0][1])
    integral = 0.0
    for (s0, v0), (s1, v1) in zip(points, points[1:]):
        integral += 0.5 * (float(v0) + float(v1)) * (float(s1) - float(s0))
    return integral / covered


def _nearest_station_row(rows: list[dict[str, Any]], station_m: float) -> dict[str, Any]:
    return min(
        rows,
        key=lambda row: (
            abs(_finite_float(row.get("Station s (m)")) - float(station_m)),
            _finite_float(row.get("Station s (m)")),
        ),
    )


def _canonical_fingerprint_payload(
    *,
    summary_payload: Mapping[str, Any],
    tendon_rows: list[dict[str, Any]],
    station_rows: list[dict[str, Any]],
    member_length_m: float,
) -> dict[str, Any]:
    return {
        "schema": EFFECTIVE_PRESTRESS_FEA_HANDOFF_SCHEMA,
        "member_length_m": round(float(member_length_m), 9),
        "averaging_basis": str(summary_payload.get("averaging_basis") or ""),
        "weighted_fpj_mpa": summary_payload.get("weighted_fpj_mpa"),
        "total_aps_mm2": summary_payload.get("total_aps_mm2"),
        "average_total_loss_mpa": summary_payload.get("average_total_loss_mpa"),
        "average_effective_stress_mpa": summary_payload.get(
            "average_effective_stress_mpa"
        ),
        "average_effective_force_kn": summary_payload.get(
            "average_effective_force_kn"
        ),
        "tendon_rows": tendon_rows,
        "station_rows": station_rows,
    }


def build_effective_prestress_fea_handoff(
    summary_payload: Mapping[str, Any],
    *,
    member_length_m: float,
    generated_at_utc: str | None = None,
    audit_metadata: Mapping[str, Any] | None = None,
    application_route: str = FEA_ROUTE_DIRECT_EFFECTIVE_FORCE,
    engineer_adopted_td: bool = True,
) -> dict[str, Any]:
    """Build a deterministic, external-FEA-only Effective Prestress handoff."""

    issues: list[str] = []
    route = str(application_route or FEA_ROUTE_DIRECT_EFFECTIVE_FORCE).strip().upper()
    if route not in FEA_APPLICATION_ROUTES:
        issues.append(f"Unsupported FEA application route: {application_route}")
        route = FEA_ROUTE_DIRECT_EFFECTIVE_FORCE
    audit = dict(audit_metadata or summary_payload.get("handoff_audit_metadata") or {})
    if not bool(summary_payload.get("effective_preview_ready")):
        issues.append(
            "Effective Prestress preview is not CURRENT/closed; refresh all loss components before export."
        )
    if not bool(summary_payload.get("projected_coverage_ready")):
        issues.append("Projected station coverage is incomplete for one or more Tendons.")

    effective_rows = _records(summary_payload.get("effective_station_rows"))
    if not effective_rows:
        issues.append("No Tendon/station Effective Prestress rows are available.")

    tendon_ids = sorted(
        {
            str(row.get("Tendon") or "").strip()
            for row in effective_rows
            if str(row.get("Tendon") or "").strip()
        }
    )
    tendon_handoff_rows: list[dict[str, Any]] = []
    station_handoff_rows: list[dict[str, Any]] = []
    tolerance = max(1.0e-7 * max(float(member_length_m), 1.0), 1.0e-9)

    for tendon_id in tendon_ids:
        tendon_source = [
            row for row in effective_rows if str(row.get("Tendon") or "") == tendon_id
        ]
        collapsed = _collapse_tendon_station_rows(tendon_source)
        if not collapsed:
            issues.append(f"{tendon_id}: no usable station rows.")
            continue
        if (
            abs(_finite_float(collapsed[0].get("Station s (m)"))) > tolerance
            or abs(
                _finite_float(collapsed[-1].get("Station s (m)"))
                - float(member_length_m)
            )
            > tolerance
        ):
            issues.append(
                f"{tendon_id}: station coverage must include s=0 and s={float(member_length_m):.3f} m."
            )

        average_fpe = _trapezoidal_average(collapsed, "fpe preview (MPa)")
        average_loss = _trapezoidal_average(collapsed, "Total loss (MPa)")
        average_fpj = _trapezoidal_average(collapsed, "fpj (MPa)")
        aps = _finite_float(collapsed[0].get("Aps (mm²)"))
        if average_fpe is None or average_loss is None or average_fpj is None or aps <= 0.0:
            issues.append(f"{tendon_id}: incomplete average stress/area source.")
            continue

        left = collapsed[0]
        mid = _nearest_station_row(collapsed, 0.5 * float(member_length_m))
        right = collapsed[-1]
        average_pe = aps * average_fpe / 1000.0
        average_pj = aps * average_fpj / 1000.0
        average_loss_percent = 100.0 * average_loss / average_fpj if average_fpj > 0 else 0.0
        remaining_percent = 100.0 * average_fpe / average_fpj if average_fpj > 0 else 0.0

        tendon_handoff_rows.append(
            {
                "Tendon": tendon_id,
                "Aps (mm²)": aps,
                "fpj (MPa)": average_fpj,
                "Pj (kN)": average_pj,
                "Left s (m)": _finite_float(left.get("Station s (m)")),
                "Left fpe (MPa)": _finite_float(left.get("fpe preview (MPa)")),
                "Left Pe (kN)": _finite_float(left.get("Pe preview (kN)")),
                "Mid s (m)": _finite_float(mid.get("Station s (m)")),
                "Mid fpe (MPa)": _finite_float(mid.get("fpe preview (MPa)")),
                "Mid Pe (kN)": _finite_float(mid.get("Pe preview (kN)")),
                "Right s (m)": _finite_float(right.get("Station s (m)")),
                "Right fpe (MPa)": _finite_float(right.get("fpe preview (MPa)")),
                "Right Pe (kN)": _finite_float(right.get("Pe preview (kN)")),
                "Average fpe (MPa)": average_fpe,
                "Average Pe (kN)": average_pe,
                "Average total loss (MPa)": average_loss,
                "Average loss (% fpj)": average_loss_percent,
                "Remaining prestress (%)": remaining_percent,
            }
        )

        profile_points = [("Left", left), ("Mid", mid), ("Right", right)]
        seen_profile_stations: set[float] = set()
        for profile_point, row in profile_points:
            profile_station = round(_finite_float(row.get("Station s (m)")), 9)
            if profile_station in seen_profile_stations:
                continue
            seen_profile_stations.add(profile_station)
            station_handoff_rows.append(
                {
                    "Tendon": tendon_id,
                    "Profile point": profile_point,
                    "Station s (m)": profile_station,
                    "Point / face source": str(row.get("Point") or ""),
                    "Aps (mm²)": aps,
                    "fpj (MPa)": _finite_float(row.get("fpj (MPa)")),
                    "Pj (kN)": _finite_float(row.get("Pj (kN)")),
                    "Friction loss (MPa)": _finite_float(row.get("Friction (MPa)")),
                    "Anchorage loss (MPa)": _finite_float(row.get("Anchorage (MPa)")),
                    "Elastic-shortening loss (MPa)": _finite_float(row.get("ES (MPa)")),
                    "Creep loss (MPa)": _finite_float(row.get("Creep (MPa)")),
                    "Shrinkage loss (MPa)": _finite_float(row.get("Shrinkage (MPa)")),
                    "Relaxation loss (MPa)": _finite_float(row.get("Relaxation (MPa)")),
                    "Total loss (MPa)": _finite_float(row.get("Total loss (MPa)")),
                    "Loss (% fpj)": _finite_float(row.get("Loss (% fpj)")),
                    "fpe (MPa)": _finite_float(row.get("fpe preview (MPa)")),
                    "Pe (kN)": _finite_float(row.get("Pe preview (kN)")),
                    "Source rows collapsed": int(row.get("Source rows collapsed") or 1),
                }
            )

    fingerprint_payload = _canonical_fingerprint_payload(
        summary_payload=summary_payload,
        tendon_rows=tendon_handoff_rows,
        station_rows=station_handoff_rows,
        member_length_m=float(member_length_m),
    )
    fingerprint = hashlib.sha256(
        json.dumps(
            fingerprint_payload,
            sort_keys=True,
            ensure_ascii=False,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
    ).hexdigest()

    generated = generated_at_utc or datetime.now(timezone.utc).isoformat(timespec="seconds")
    source_ready = bool(not issues and tendon_handoff_rows and station_handoff_rows)
    # A closed, validated loss chain is the engineering source gate.  No
    # additional ceremonial checkbox is required to export the same result.
    download_ready = source_ready
    if not source_ready:
        status = "SOURCE BLOCKED"
    else:
        status = AUTOMATIC_HANDOFF_READY_STATUS

    source_id = fingerprint[:12]
    contract_payload = {
        "schema": EFFECTIVE_PRESTRESS_FEA_HANDOFF_SCHEMA,
        "source_fingerprint": fingerprint,
        "application_route": route,
        "automatic_source_adoption": bool(source_ready),
    }
    contract_fingerprint = hashlib.sha256(
        json.dumps(contract_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    contract_id = contract_fingerprint[:16]

    for row in tendon_handoff_rows:
        row["Source ID"] = source_id
    for row in station_handoff_rows:
        row["Source ID"] = source_id
    system_station_rows = _records(summary_payload.get("system_station_rows"))
    for row in system_station_rows:
        row["Source ID"] = source_id

    def add_summary(rows: list[dict[str, Any]], field: str, value: Any) -> None:
        if value is None:
            return
        if isinstance(value, str) and not value.strip():
            return
        rows.append({"Field": field, "Value": value})

    event_schedule = audit.get("permanent_load_event_schedule")
    if isinstance(event_schedule, Sequence) and not isinstance(event_schedule, (str, bytes, bytearray)):
        event_parts: list[str] = []
        for item in event_schedule:
            if not isinstance(item, Mapping):
                continue
            event_id = str(item.get("Event") or item.get("Event ID") or "PL").strip()
            group = str(item.get("Permanent load group") or item.get("Load group") or "Permanent load").strip()
            age = item.get("Age (days)", item.get("Activation age (days)"))
            case = str(item.get("Imported FEA case") or item.get("Case Name") or "").strip()
            age_text = f" at {float(age):.1f} d" if age is not None else ""
            case_text = f" [{case}]" if case else ""
            event_parts.append(f"{event_id}: {group}{age_text}{case_text}")
        event_schedule = "; ".join(event_parts) if event_parts else "None"

    summary_rows: list[dict[str, Any]] = []
    add_summary(summary_rows, "Schema version", EFFECTIVE_PRESTRESS_FEA_HANDOFF_SCHEMA)
    add_summary(summary_rows, "Handoff status", status)
    add_summary(summary_rows, "Source adoption", "AUTOMATIC WHEN CURRENT/CLOSED")
    add_summary(summary_rows, "FEA application route", route)
    add_summary(summary_rows, "Generated at (UTC)", generated)
    add_summary(summary_rows, "Source ID", source_id)
    add_summary(summary_rows, "Source fingerprint", fingerprint)
    add_summary(summary_rows, "Contract ID", contract_id)
    add_summary(summary_rows, "Project name", audit.get("project_name"))
    add_summary(summary_rows, "Member ID", audit.get("member_id"))
    add_summary(summary_rows, "Member workflow", audit.get("member_workflow"))
    add_summary(summary_rows, "Construction type", audit.get("construction_type"))
    add_summary(summary_rows, "Member design code", audit.get("member_design_code"))
    add_summary(summary_rows, "Prestress-loss basis", audit.get("prestress_loss_basis") or "AASHTO LRFD 2020 §5.4.2.3 + §5.9.3.4")
    add_summary(summary_rows, "Member length (m)", float(member_length_m))
    add_summary(summary_rows, "Tendons represented", len(tendon_handoff_rows))
    add_summary(summary_rows, "Three-point profile rows", len(station_handoff_rows))
    add_summary(summary_rows, "Total Aps (mm²)", summary_payload.get("total_aps_mm2"))
    add_summary(summary_rows, "Average fpj (MPa)", summary_payload.get("weighted_fpj_mpa"))
    add_summary(summary_rows, "Average total loss (MPa)", summary_payload.get("average_total_loss_mpa"))
    add_summary(summary_rows, "Average total loss (% fpj)", summary_payload.get("average_total_loss_percent"))
    add_summary(summary_rows, "Average fpe (MPa)", summary_payload.get("average_effective_stress_mpa"))
    add_summary(summary_rows, "Initial total tendon force (kN)", summary_payload.get("initial_total_force_kn"))
    add_summary(summary_rows, "Average effective tendon force (kN)", summary_payload.get("average_effective_force_kn"))
    add_summary(summary_rows, "Stressing / jacking mode", audit.get("stressing_mode"))
    add_summary(summary_rows, "Tendon stressing age ti (days)", audit.get("ti_days"))
    add_summary(summary_rows, "Tendon grouting age tg (days)", audit.get("tg_days"))
    add_summary(summary_rows, "Falsework removal age tr (days)", audit.get("tr_days"))
    add_summary(summary_rows, "Final age tf (days)", audit.get("tf_days"))
    add_summary(summary_rows, "Permanent-load events", event_schedule if event_schedule is not None else "None")
    add_summary(summary_rows, "Relative humidity RH (%)", audit.get("rh_percent"))
    add_summary(summary_rows, "Concrete f'c (MPa)", audit.get("fc_mpa"))
    add_summary(summary_rows, "Stressing strength f'ci (MPa)", audit.get("fci_mpa"))
    add_summary(summary_rows, "Member-equivalent V/S (mm)", audit.get("v_over_s_mm"))
    add_summary(summary_rows, "TD subtotal (MPa)", summary_payload.get("time_dependent_loss_mpa"))
    add_summary(summary_rows, "Averaging basis", EFFECTIVE_PRESTRESS_FEA_HANDOFF_BASIS)
    add_summary(summary_rows, "TD loss basis", "Validated system-average total accounted loss (% fpj)")
    add_summary(summary_rows, "Secondary prestress", "Not included — calculate from structural restraint in external FEA")
    add_summary(summary_rows, "SLS return route", "Import verified FEA SLS P/V2/M3 responses in the main Loads workspace")

    route_instruction = (
        "Input exported effective fpe/Pe directly and disable all duplicate FEA loss calculations."
        if route == FEA_ROUTE_DIRECT_EFFECTIVE_FORCE
        else "Input fpj/Pj and reproduce the same station loss profile in FEA; do not also apply exported fpe/Pe."
    )
    instructions_rows = [
        {
            "Topic": "Purpose",
            "Instruction": "This handoff transfers the CURRENT/CLOSED loss result to external FEA. Source readiness is automatic after validation; no separate confirmation is required.",
        },
        {
            "Topic": "Adopted application route",
            "Instruction": f"{route}: {route_instruction}",
        },
        {
            "Topic": "Units and sign",
            "Instruction": "Station s is metres from the left member end. Strand stress is MPa tension-positive; Pe/Pj are positive tendon tension magnitudes in kN.",
        },
        {
            "Topic": "No double counting",
            "Instruction": "Use exactly one application route. Never apply exported effective fpe/Pe and then calculate friction, anchorage, ES, creep, shrinkage, or relaxation again in FEA.",
        },
        {
            "Topic": "Three-point force profile",
            "Instruction": "The exported profile contains Left / Mid / Right values only. It is a compact three-point review profile, not a full tendon-station or tendon-arc-length profile.",
        },
        {
            "Topic": "Secondary prestress",
            "Instruction": "Secondary prestress is not a tendon loss and must not be subtracted from Pe. Let the external portal-frame model calculate primary and secondary response from compatibility/restraint.",
        },
        {
            "Topic": "Time-dependent limitation",
            "Instruction": "The system-average total accounted loss is expressed as %fpj. Subtract its MPa equivalent once from the pre-loss FEA tendon-stress basis, then apply the resulting force to each tendon.",
        },
        {
            "Topic": "SLS feedback",
            "Instruction": "After FEA calculates primary + secondary prestress and service actions, export verified SLS P/V2/M3 responses and import them through the main Loads workspace.",
        },
        {
            "Topic": "Traceability",
            "Instruction": "Record the Source ID, full source fingerprint, Contract ID, application route, and adoption statement in the FEA model note/report. Regenerate the handoff whenever source inputs change.",
        },
    ]

    return {
        "ready": source_ready,
        "download_ready": download_ready,
        "status": status,
        "issues": issues,
        "schema_version": EFFECTIVE_PRESTRESS_FEA_HANDOFF_SCHEMA,
        "source_id": source_id,
        "source_fingerprint": fingerprint,
        "contract_id": contract_id,
        "contract_fingerprint": contract_fingerprint,
        "generated_at_utc": generated,
        "application_route": route,
        "engineer_adopted_td": bool(source_ready),
        "summary_rows": summary_rows,
        "tendon_rows": tendon_handoff_rows,
        "station_rows": station_handoff_rows,
        "system_station_rows": system_station_rows,
        "instructions_rows": instructions_rows,
        "scope_guard": (
            "This is a preliminary external-FEA contract using a representative TD-loss approximation. External FEA must calculate portal-frame secondary prestress. Concrete Section Pro does not automatically feed this handoff into SLS; verified FEA SLS responses must return through Loads."
        ),
    }


def _header_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value): int(cell.column)
        for cell in worksheet[1]
        if cell.value is not None
    }


def _excel_ref(sheet_name: str, column_index: int, row_index: int) -> str:
    quoted = sheet_name.replace("'", "''")
    return f"'{quoted}'!{get_column_letter(column_index)}{row_index}"


def _summary_value_ref(worksheet, field_name: str) -> str:
    for row in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row, column=1).value or "") == field_name:
            return _excel_ref(worksheet.title, 2, row)
    raise KeyError(field_name)


def _write_qa_checks(workbook) -> None:
    qa = workbook["QA Checks"]
    headers = [
        "Scope",
        "Source row",
        "Stress closure (MPa)",
        "Force closure (kN)",
        "Tolerance",
        "Status",
        "Stress check basis",
        "Force check basis",
    ]
    for col, value in enumerate(headers, start=1):
        qa.cell(row=1, column=col, value=value)
    row_out = 2

    summary = workbook["Handoff Summary"]
    qa.cell(row=row_out, column=1, value="System average")
    qa.cell(row=row_out, column=2, value="Handoff Summary")
    qa.cell(
        row=row_out,
        column=3,
        value=f"={_summary_value_ref(summary, 'Average fpj (MPa)')}-{_summary_value_ref(summary, 'Average total loss (MPa)')}-{_summary_value_ref(summary, 'Average fpe (MPa)')}",
    )
    qa.cell(
        row=row_out,
        column=4,
        value=f"={_summary_value_ref(summary, 'Initial total tendon force (kN)')}-{_summary_value_ref(summary, 'Average effective tendon force (kN)')}-{_summary_value_ref(summary, 'Total Aps (mm²)')}*{_summary_value_ref(summary, 'Average total loss (MPa)')}/1000",
    )
    qa.cell(row=row_out, column=5, value=1.0e-6)
    qa.cell(row=row_out, column=6, value=f'=IF(AND(ABS(C{row_out})<=E{row_out},ABS(D{row_out})<=E{row_out}),"PASS","REVIEW")')
    qa.cell(row=row_out, column=7, value="Average fpj − average loss − average fpe")
    qa.cell(row=row_out, column=8, value="ΣPj − ΣPe,avg − ApsΣ·average loss/1000")
    row_out += 1

    tendon_ws = workbook["Tendon Handoff"]
    th = _header_map(tendon_ws)
    for source_row in range(2, tendon_ws.max_row + 1):
        qa.cell(row=row_out, column=1, value="Tendon average")
        qa.cell(row=row_out, column=2, value=f"Tendon Handoff row {source_row}")
        qa.cell(
            row=row_out,
            column=3,
            value=f"={_excel_ref(tendon_ws.title, th['fpj (MPa)'], source_row)}-{_excel_ref(tendon_ws.title, th['Average total loss (MPa)'], source_row)}-{_excel_ref(tendon_ws.title, th['Average fpe (MPa)'], source_row)}",
        )
        qa.cell(
            row=row_out,
            column=4,
            value=f"={_excel_ref(tendon_ws.title, th['Pj (kN)'], source_row)}-{_excel_ref(tendon_ws.title, th['Average Pe (kN)'], source_row)}-{_excel_ref(tendon_ws.title, th['Aps (mm²)'], source_row)}*{_excel_ref(tendon_ws.title, th['Average total loss (MPa)'], source_row)}/1000",
        )
        qa.cell(row=row_out, column=5, value=1.0e-6)
        qa.cell(row=row_out, column=6, value=f'=IF(AND(ABS(C{row_out})<=E{row_out},ABS(D{row_out})<=E{row_out}),"PASS","REVIEW")')
        qa.cell(row=row_out, column=7, value="fpj − average total loss − average fpe")
        qa.cell(row=row_out, column=8, value="Pj − average Pe − Aps·average loss/1000")
        row_out += 1

    profile_ws = workbook["Three-Point Profile"]
    ph = _header_map(profile_ws)
    loss_headers = [
        "Friction loss (MPa)",
        "Anchorage loss (MPa)",
        "Elastic-shortening loss (MPa)",
        "Creep loss (MPa)",
        "Shrinkage loss (MPa)",
        "Relaxation loss (MPa)",
    ]
    for source_row in range(2, profile_ws.max_row + 1):
        qa.cell(row=row_out, column=1, value="Three-point profile")
        qa.cell(row=row_out, column=2, value=f"Three-Point Profile row {source_row}")
        loss_refs = "+".join(_excel_ref(profile_ws.title, ph[name], source_row) for name in loss_headers)
        qa.cell(
            row=row_out,
            column=3,
            value=f"={_excel_ref(profile_ws.title, ph['fpj (MPa)'], source_row)}-({loss_refs})-{_excel_ref(profile_ws.title, ph['fpe (MPa)'], source_row)}",
        )
        qa.cell(
            row=row_out,
            column=4,
            value=f"={_excel_ref(profile_ws.title, ph['Pj (kN)'], source_row)}-{_excel_ref(profile_ws.title, ph['Pe (kN)'], source_row)}-{_excel_ref(profile_ws.title, ph['Aps (mm²)'], source_row)}*{_excel_ref(profile_ws.title, ph['Total loss (MPa)'], source_row)}/1000",
        )
        qa.cell(row=row_out, column=5, value=1.0e-6)
        qa.cell(row=row_out, column=6, value=f'=IF(AND(ABS(C{row_out})<=E{row_out},ABS(D{row_out})<=E{row_out}),"PASS","REVIEW")')
        qa.cell(row=row_out, column=7, value="fpj − Σ(component losses) − fpe")
        qa.cell(row=row_out, column=8, value="Pj − Pe − Aps·total loss/1000")
        row_out += 1


def _style_export_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    if worksheet.title == "Three-Point Profile":
        worksheet.freeze_panes = "C2"
    elif worksheet.title in {"Tendon Handoff", "System Station", "QA Checks"}:
        worksheet.freeze_panes = "B2"
    else:
        worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    if worksheet.title == "Handoff Summary":
        for row in range(2, worksheet.max_row + 1):
            field = str(worksheet.cell(row=row, column=1).value or "")
            value_cell = worksheet.cell(row=row, column=2)
            if "%" in field:
                value_cell.number_format = "0.000"
            elif any(token in field for token in ("(MPa)", "(kN)", "(m)", "(mm²)", "(days)")):
                value_cell.number_format = "0.000"
            elif field in {"Tendons represented", "Three-point profile rows"}:
                value_cell.number_format = "0"
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)

    header_map = _header_map(worksheet)
    for header, col in header_map.items():
        col_letter = get_column_letter(col)
        text = str(header)
        if worksheet.title == "Handoff Summary":
            worksheet.column_dimensions["A"].width = 38
            worksheet.column_dimensions["B"].width = 72
            break
        if worksheet.title == "Instructions":
            worksheet.column_dimensions["A"].width = 28
            worksheet.column_dimensions["B"].width = 95
            break
        if text in {"Tendon", "Source ID", "Status", "Scope"}:
            width = 16
        elif "Instruction" in text or "basis" in text.lower():
            width = 42
        elif "Point" in text or "Source row" in text:
            width = 24
        else:
            width = min(max(len(text) + 2, 12), 22)
        worksheet.column_dimensions[col_letter].width = width

        if "closure" in text.lower():
            number_format = "0.000E+00"
        elif "%" in text:
            number_format = "0.000"
        elif any(token in text for token in ("(MPa)", "(kN)", "(m)", "(mm²)")):
            number_format = "0.000"
        else:
            number_format = None
        if number_format:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right", vertical="top")

    if worksheet.title == "QA Checks":
        for row in range(2, worksheet.max_row + 1):
            status_cell = worksheet.cell(row=row, column=6)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 32


def effective_prestress_handoff_excel_bytes(handoff: Mapping[str, Any]) -> bytes:
    """Return a compact, formula-audited multi-sheet XLSX handoff workbook."""

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(_records(handoff.get("summary_rows"))).to_excel(
            writer, sheet_name="Handoff Summary", index=False
        )
        tendon_df = pd.DataFrame(_records(handoff.get("tendon_rows")))
        if not tendon_df.empty:
            tendon_df = tendon_df[[column for column in TENDON_HANDOFF_EXPORT_COLUMNS if column in tendon_df.columns]]
        tendon_df.to_excel(writer, sheet_name="Tendon Handoff", index=False)
        pd.DataFrame(_records(handoff.get("station_rows"))).to_excel(
            writer, sheet_name="Three-Point Profile", index=False
        )
        pd.DataFrame(_records(handoff.get("system_station_rows"))).to_excel(
            writer, sheet_name="System Station", index=False
        )
        pd.DataFrame(columns=["Scope", "Source row", "Stress closure (MPa)", "Force closure (kN)", "Tolerance", "Status", "Stress check basis", "Force check basis"]).to_excel(
            writer, sheet_name="QA Checks", index=False
        )
        pd.DataFrame(_records(handoff.get("instructions_rows"))).to_excel(
            writer, sheet_name="Instructions", index=False
        )
        _write_qa_checks(writer.book)
        for worksheet in writer.book.worksheets:
            _style_export_sheet(worksheet)
        calculation = getattr(writer.book, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
            calculation.calcMode = "auto"
    return output.getvalue()


def effective_prestress_handoff_csv_bytes(
    handoff: Mapping[str, Any], *, table: str
) -> bytes:
    """Return UTF-8-SIG CSV bytes for the selected handoff table."""

    table_map = {
        "tendon": "tendon_rows",
        "station": "station_rows",
        "profile": "station_rows",
        "three-point": "station_rows",
        "system": "system_station_rows",
    }
    key = table_map.get(str(table).casefold())
    if key is None:
        raise ValueError(f"Unsupported handoff CSV table: {table}")
    dataframe = pd.DataFrame(_records(handoff.get(key)))
    if str(table).casefold() == "tendon" and not dataframe.empty:
        dataframe = dataframe[[column for column in TENDON_HANDOFF_EXPORT_COLUMNS if column in dataframe.columns]]
    return dataframe.to_csv(index=False).encode("utf-8-sig")
