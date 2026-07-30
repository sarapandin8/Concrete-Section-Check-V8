from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from concrete_pmm_pro.crossbeam.effective_prestress_handoff import (
    build_effective_prestress_fea_handoff,
    effective_prestress_handoff_csv_bytes,
    effective_prestress_handoff_excel_bytes,
)


def _summary_payload() -> dict[str, object]:
    rows: list[dict[str, object]] = []
    for tendon, aps, offset in (("T1", 1000.0, 0.0), ("T2", 2000.0, 20.0)):
        for station, fpe in ((0.0, 1100.0 + offset), (10.0, 1200.0 + offset), (20.0, 1100.0 + offset)):
            fpj = 1400.0
            total_loss = fpj - fpe
            rows.append(
                {
                    "Tendon": tendon,
                    "Station s (m)": station,
                    "Point": f"P{int(station)}",
                    "Aps (mm²)": aps,
                    "fpj (MPa)": fpj,
                    "Friction (MPa)": 10.0,
                    "Anchorage (MPa)": 20.0,
                    "ES (MPa)": total_loss - 75.0,
                    "Creep (MPa)": 30.0,
                    "Shrinkage (MPa)": 10.0,
                    "Relaxation (MPa)": 5.0,
                    "TD total (MPa)": 45.0,
                    "Total loss (MPa)": total_loss,
                    "Loss (% fpj)": 100.0 * total_loss / fpj,
                    "fpe preview (MPa)": fpe,
                    "Pj (kN)": aps * fpj / 1000.0,
                    "Pe preview (kN)": aps * fpe / 1000.0,
                }
            )
    average_fpe = (1000.0 * 1150.0 + 2000.0 * 1170.0) / 3000.0
    return {
        "effective_preview_ready": True,
        "projected_coverage_ready": True,
        "member_length_m": 20.0,
        "weighted_fpj_mpa": 1400.0,
        "total_aps_mm2": 3000.0,
        "average_total_loss_mpa": 1400.0 - average_fpe,
        "average_total_loss_percent": 100.0 * (1400.0 - average_fpe) / 1400.0,
        "average_effective_stress_mpa": average_fpe,
        "initial_total_force_kn": 4200.0,
        "average_effective_force_kn": 3000.0 * average_fpe / 1000.0,
        "averaging_basis": "Aps-weighted projected-station trapezoidal average",
        "effective_station_rows": rows,
        "system_station_rows": [],
    }


def test_ptloss4d2_handoff_exports_left_mid_right_and_projected_average() -> None:
    handoff = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
        engineer_adopted_td=True,
    )

    assert handoff["ready"] is True
    assert handoff["download_ready"] is True
    assert handoff["status"] == "ENGINEER-ADOPTED PRELIMINARY HANDOFF — EXTERNAL FEA ONLY"
    assert len(handoff["tendon_rows"]) == 2
    t1 = handoff["tendon_rows"][0]
    assert t1["Tendon"] == "T1"
    assert t1["Left Pe (kN)"] == pytest.approx(1100.0)
    assert t1["Mid Pe (kN)"] == pytest.approx(1200.0)
    assert t1["Right Pe (kN)"] == pytest.approx(1100.0)
    assert t1["Average fpe (MPa)"] == pytest.approx(1150.0)
    assert t1["Average Pe (kN)"] == pytest.approx(1150.0)
    assert t1["Average total loss (MPa)"] == pytest.approx(250.0)
    assert t1["Remaining prestress (%)"] == pytest.approx(100.0 * 1150.0 / 1400.0)
    assert len(handoff["station_rows"]) == 6
    assert all(row["Source ID"] == handoff["source_id"] for row in handoff["station_rows"])
    assert all("Source fingerprint" not in row for row in handoff["station_rows"])


def test_ptloss4d2_fingerprint_is_independent_of_export_timestamp() -> None:
    first = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
    )
    second = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-30T12:00:00+00:00",
    )
    assert first["source_fingerprint"] == second["source_fingerprint"]


def test_ptloss4d2_workbook_and_csv_contain_auditable_handoff_tables() -> None:
    handoff = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
    )
    workbook_bytes = effective_prestress_handoff_excel_bytes(handoff)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook.sheetnames == [
        "Handoff Summary",
        "Tendon Handoff",
        "Three-Point Profile",
        "System Station",
        "QA Checks",
        "Instructions",
    ]
    assert workbook["Tendon Handoff"]["A2"].value == "T1"
    assert "Secondary prestress" in {
        workbook["Instructions"].cell(row=row, column=1).value
        for row in range(2, workbook["Instructions"].max_row + 1)
    }
    qa = workbook["QA Checks"]
    assert str(qa["C2"].value).startswith("=")
    assert str(qa["D2"].value).startswith("=")
    assert str(qa["F2"].value).startswith("=IF")
    summary_values = {
        workbook["Handoff Summary"].cell(row=row, column=1).value:
        workbook["Handoff Summary"].cell(row=row, column=2).value
        for row in range(2, workbook["Handoff Summary"].max_row + 1)
    }
    assert summary_values["FEA application route"] == "DIRECT_EFFECTIVE_FORCE"
    assert summary_values["Engineer TD adoption"] == "REQUIRED"
    assert "Source fingerprint" not in {cell.value for cell in workbook["Tendon Handoff"][1]}
    assert "Source ID" in {cell.value for cell in workbook["Tendon Handoff"][1]}
    tendon_csv = effective_prestress_handoff_csv_bytes(handoff, table="tendon")
    profile_csv = effective_prestress_handoff_csv_bytes(handoff, table="profile")
    assert b"Average Pe (kN)" in tendon_csv
    assert b"Station s (m)" in profile_csv


def test_ptloss4d2_handoff_blocks_unclosed_or_incomplete_sources() -> None:
    payload = _summary_payload()
    payload["effective_preview_ready"] = False
    payload["projected_coverage_ready"] = False
    handoff = build_effective_prestress_fea_handoff(
        payload,
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
    )
    assert handoff["ready"] is False
    assert handoff["download_ready"] is False
    assert handoff["status"] == "SOURCE BLOCKED"
    assert len(handoff["issues"]) >= 2


def test_ptloss4d2_ui_keeps_external_fea_and_sls_return_boundaries_explicit() -> None:
    source = Path("concrete_pmm_pro/ui/crossbeam_pages.py").read_text(encoding="utf-8")
    assert "FEA Effective Prestress Handoff" in source
    assert "Download FEA handoff workbook" in source
    assert "Do not apply the same losses twice" in source
    assert "Engineer adoption — use the representative Time-Dependent loss approximation" in source
    assert "DIRECT_EFFECTIVE_FORCE" in source
    assert "Download Three-Point Profile CSV" in source
    assert "disabled=not download_ready" in source
    assert "calculate secondary; return verified P/V2/M3 through Loads" in source


def test_ptloss4d2a_requires_explicit_td_adoption_before_download() -> None:
    source_ready = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
        engineer_adopted_td=False,
    )
    assert source_ready["ready"] is True
    assert source_ready["download_ready"] is False
    assert source_ready["status"] == "PRELIMINARY SOURCE READY — ENGINEER ADOPTION REQUIRED"

    adopted = build_effective_prestress_fea_handoff(
        _summary_payload(),
        member_length_m=20.0,
        generated_at_utc="2026-07-29T12:00:00+00:00",
        engineer_adopted_td=True,
    )
    assert adopted["download_ready"] is True
    assert adopted["source_fingerprint"] == source_ready["source_fingerprint"]
    assert adopted["contract_id"] != source_ready["contract_id"]


def test_ptloss4d2a_profile_contract_exports_exactly_left_mid_right() -> None:
    payload = _summary_payload()
    extra_rows: list[dict[str, object]] = []
    for row in list(payload["effective_station_rows"]):
        if float(row["Station s (m)"]) == 0.0:
            extra = dict(row)
            extra["Station s (m)"] = 5.0
            extra["Point"] = "P5"
            extra["fpe preview (MPa)"] = float(row["fpe preview (MPa)"]) + 50.0
            extra["Total loss (MPa)"] = float(row["fpj (MPa)"]) - float(extra["fpe preview (MPa)"])
            extra["Pe preview (kN)"] = float(row["Aps (mm²)"]) * float(extra["fpe preview (MPa)"]) / 1000.0
            extra_rows.append(extra)
    payload["effective_station_rows"] = list(payload["effective_station_rows"]) + extra_rows
    handoff = build_effective_prestress_fea_handoff(
        payload,
        member_length_m=20.0,
        engineer_adopted_td=True,
    )
    assert len(handoff["station_rows"]) == 6
    assert {row["Profile point"] for row in handoff["station_rows"]} == {"Left", "Mid", "Right"}
